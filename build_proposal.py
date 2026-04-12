#!/usr/bin/env python3
"""Generate VentureTown Game Proposal Excel file."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style helpers ──
TITLE_FONT = Font(name='Microsoft JhengHei', size=18, bold=True, color='FFFFFF')
H1_FONT = Font(name='Microsoft JhengHei', size=14, bold=True, color='8B4513')
H2_FONT = Font(name='Microsoft JhengHei', size=11, bold=True, color='333333')
BODY_FONT = Font(name='Microsoft JhengHei', size=10)
SMALL_FONT = Font(name='Microsoft JhengHei', size=9, color='666666')
HEADER_FILL = PatternFill('solid', fgColor='8B4513')
H1_FILL = PatternFill('solid', fgColor='FAEBD7')
TABLE_HEADER_FILL = PatternFill('solid', fgColor='DEB887')
ALT_ROW_FILL = PatternFill('solid', fgColor='FFF8F0')
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D2B48C'),
    right=Side(style='thin', color='D2B48C'),
    top=Side(style='thin', color='D2B48C'),
    bottom=Side(style='thin', color='D2B48C'),
)

def setup_sheet(ws, title, col_widths):
    ws.title = title
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)] = ws.column_dimensions[get_column_letter(i)]
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_properties.tabColor = '8B4513'

def add_title_row(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    ws.row_dimensions[row].height = 40

def add_section(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = H1_FONT
    c.fill = H1_FILL
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 28
    return row + 1

def add_table_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name='Microsoft JhengHei', size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='A0522D')
        c.alignment = CENTER
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 22
    return row + 1

def add_table_row(ws, row, values, alt=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY_FONT
        c.alignment = WRAP
        c.border = THIN_BORDER
        if alt:
            c.fill = ALT_ROW_FILL
    return row + 1

def add_text(ws, row, col, text, font=None, cols_merge=0):
    if cols_merge > 0:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+cols_merge-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font or BODY_FONT
    c.alignment = WRAP
    return row + 1

# ══════════════════════════════════════
# Sheet 1: 遊戲概述與流程
# ══════════════════════════════════════
ws = wb.active
setup_sheet(ws, '遊戲概述', [4, 20, 30, 25])
r = 1
add_title_row(ws, r, '創業小鎮（VentureTown）— 遊戲企劃書', 4); r += 2

r = add_section(ws, r, '一、遊戲概述', 4)
r = add_text(ws, r, 1, '遊戲名稱', H2_FONT)
ws.cell(row=r-1, column=2, value='創業小鎮（VentureTown）').font = BODY_FONT
r = add_text(ws, r, 1, '類型', H2_FONT)
ws.cell(row=r-1, column=2, value='網頁策略經營遊戲').font = BODY_FONT
r = add_text(ws, r, 1, '平台', H2_FONT)
ws.cell(row=r-1, column=2, value='瀏覽器（單檔 HTML）').font = BODY_FONT
r = add_text(ws, r, 1, '核心概念', H2_FONT)
ws.cell(row=r-1, column=2, value='在 4×4 格子小鎮上放置設施，投入資源讓其沿路徑經過設施轉換增值，以金錢形式產出收益，達成目標即過關。').font = BODY_FONT
ws.merge_cells(start_row=r-1, start_column=2, end_row=r-1, end_column=4)
r += 1

r = add_section(ws, r, '二、資源系統', 4)
r = add_table_header(ws, r, ['圖示', '資源名稱', '說明', '備註'])
r = add_table_row(ws, r, ['💰', '金錢 (Money)', '基礎資源，投入時的初始類型', '初始值 = 1'])
r = add_table_row(ws, r, ['🪨', '原料 (Material)', '由金錢轉換而來', '原料廠產出'], True)
r = add_table_row(ws, r, ['📦', '商品 (Goods)', '由原料加工而來', '工廠產出'])
r = add_table_row(ws, r, ['💎', '鑽石 (Diamond)', '特殊資源', '嫉妒惡魔專屬'], True)
r += 1

r = add_section(ws, r, '三、回合制流程', 4)
flow_steps = [
    ('1', '回合開始', '系統隨機提供 2 個商業行動選項', '每輪 10 回合'),
    ('2', '選擇行動', '玩家從行動中二選一（或跳過）', '行動費用從收益扣除'),
    ('3', '放置設施', '從手牌將設施拖曳到空格放置', '可放多個'),
    ('4', '投入資源', '將資源卡拖曳到地圖邊緣投入', '資源沿方向前進'),
    ('5', '設施轉換', '資源經過設施自動觸發轉換與增值', '金錢→原料→商品→金錢'),
    ('6', '收益結算', '資源離開時：最終金錢值 - 初始值(1) = 收益', '累加至總收益'),
    ('7', '結束回合', '點擊「結束回合」進入下回合', '每 3 回合觸發市場事件'),
]
r = add_table_header(ws, r, ['步驟', '階段', '說明', '備註'])
for i, s in enumerate(flow_steps):
    r = add_table_row(ws, r, s, i % 2 == 1)
r += 1

r = add_section(ws, r, '四、勝敗條件與難度', 4)
r = add_text(ws, r, 1, '過關條件', H2_FONT)
ws.cell(row=r-1, column=2, value='累計收益 ≥ 當輪目標（第 1 輪目標 = 30）').font = BODY_FONT
ws.merge_cells(start_row=r-1, start_column=2, end_row=r-1, end_column=4)
r = add_text(ws, r, 1, '失敗條件', H2_FONT)
ws.cell(row=r-1, column=2, value='10 回合用盡仍未達目標').font = BODY_FONT
r = add_text(ws, r, 1, '下一輪目標', H2_FONT)
ws.cell(row=r-1, column=2, value='round(舊目標 × (1.8 × 難度倍率) + 15)').font = BODY_FONT
r = add_text(ws, r, 1, '難度倍率', H2_FONT)
ws.cell(row=r-1, column=2, value='根據玩家達標速度自動調整 0.6~2.5 倍，失敗重開時重置').font = BODY_FONT
ws.merge_cells(start_row=r-1, start_column=2, end_row=r-1, end_column=4)
r += 1

r = add_section(ws, r, '五、難度調整規則', 4)
r = add_table_header(ws, r, ['', '條件', '倍率變化', ''])
diff_rules = [
    ('', '平均 ≤2 回合達標', '+25%', ''),
    ('', '平均 ≤4 回合達標', '+15%', ''),
    ('', '平均 ≤6 回合達標', '+5%', ''),
    ('', '平均 ≤8 回合達標', '不變', ''),
    ('', '平均 >8 回合達標', '-8%', ''),
    ('', '連續 2 輪 ≤3 回合瞬殺', '額外 +10%', ''),
    ('', '連續 3 輪 ≤3 回合瞬殺', '額外 +20%', ''),
    ('', '收益 ≥2 倍目標', '額外 +8%', ''),
    ('', '收益 ≥3 倍目標', '額外 +15%', ''),
]
for i, d in enumerate(diff_rules):
    r = add_table_row(ws, r, d, i % 2 == 1)

# ══════════════════════════════════════
# Sheet 2: 設施一覽
# ══════════════════════════════════════
ws2 = wb.create_sheet()
setup_sheet(ws2, '設施一覽', [6, 12, 12, 35, 15])
r = 1
add_title_row(ws2, r, '設施一覽表', 5); r += 2

categories = [
    ('基礎設施', [
        ('⛏️', '原料廠', 'money→material', '金錢→原料 +1', '基礎'),
        ('🏭', '工廠', 'material→goods', '原料→商品 +1', '基礎'),
        ('🏪', '商店', 'goods→money', '商品→金錢 +1（isShop）', '基礎'),
        ('🔩', '精煉廠', 'material→material', '原料→原料 +2', '基礎'),
        ('📦', '倉庫', 'goods→goods', '商品→商品 +2', '基礎'),
        ('⚡', '增幅器', 'any→不轉換', '任意 +2', '基礎'),
        ('🔄', '轉化器', 'any→money', '任意→金錢', '基礎'),
    ]),
    ('商店系設施（視為商店）', [
        ('🛍️', '小型販售商', 'goods/money→money', '商品→金錢+1；金錢→商品→金錢+2', '商店系'),
        ('🎫', '黃牛販子', 'goods→goods', '商品→金錢+1→商品(-10%)', '商店系'),
        ('🛒', '量販店', 'goods→goods', '商品→金錢(+50%)→商品', '商店系'),
        ('🏬', '百貨公司', 'goods→money', '需2×2商店。商品→×2金錢→+2%商品→金錢', '商店系/2×2'),
    ]),
    ('人力流設施', [
        ('🎓', '人力訓練中心', '-', '每回合開始產生 1 人材', '人力'),
        ('⚙️', '勞動轉換站', '-', '消耗 2 人材，資源×2', '人力'),
        ('🗃️', '人材倉庫', '-', '持有 5+ 人材時額外 +5', '人力'),
        ('🏦', '人才市場', 'money→money', '金錢通過時消耗 3% 獲得 2 人材', '人力'),
        ('💼', '加班辦公室', '-', '資源通過時 +人材數量一半%', '人力'),
        ('🏧', '人力銀行', 'money→money', '金錢通過時儲存 10%，下次投入釋放', '人力'),
        ('🏢', '派遣總部', 'goods→goods', '商品通過時獲得 2 人材', '人力'),
        ('📢', '集體罷工台', '-', '消耗所有人材，每個 +10（主動觸發）', '人力'),
    ]),
    ('物流中心流設施', [
        ('🔄', '螺旋物流站', '-', '每通過 1 個設施 +2', '物流'),
        ('🏁', '終點站', '-', '通過時 +(已通過設施數×2)', '物流'),
        ('📡', '物流放大器', '-', '通過物流後下一設施效果×2', '物流'),
        ('🔀', '轉運中心', '-', '放置時決定方向（可選 4 向）', '物流'),
        ('⚡', '速遞站', '-', '通過後若下一格有設施立即再觸發', '物流'),
        ('🏪', '物流倉', '-', '儲存通過時數值 50%，下次投入釋放', '物流'),
        ('📊', '環境感應站', '-', '周圍 4 格每有一個設施 +1', '物流'),
    ]),
    ('貿易流設施', [
        ('⚓', '外貿港口', 'goods→goods', '商品每 10 一階 +10%', '貿易'),
        ('📈', '匯率波動板', '-', '不參與路徑；結算時隨機 ±3~15%', '貿易'),
        ('🏛', '期貨交易所', 'goods→money', '放置時鎖定倍率，永久使用', '貿易'),
        ('🌏', '貿易特區', '-', '周圍有商店系設施時輸出×1.5', '貿易'),
        ('🏷️', '清倉拍賣場', 'goods→money', '商品清零，獲得等值×4 金錢', '貿易'),
        ('🧾', '進出口稅站', '-', '扣除 1 收益，本格效果×2', '貿易'),
        ('🗽', '自由市場', 'any→money', '任意資源→金錢（當前數值）', '貿易'),
    ]),
    ('拆遷流設施', [
        ('💣', '爆破裝置', '-', '3 回合後消滅自身+相鄰設施', '拆遷'),
        ('🏗', '建築廢料廠', '-', '通過時消滅自身，本次投入×2', '拆遷'),
        ('🛖', '臨時工棚', '-', '每移動一次 +1（最高 +5）', '拆遷'),
        ('🧲', '磁力板', '-', '每回合開始與相鄰隨機設施交換', '拆遷'),
        ('🗿', '廢墟紀念碑', '-', '只能放在廢墟上，輸出 +5', '拆遷'),
        ('🏢', '拆遷補償局', '-', '每次移動/消滅設施 +1 收益', '拆遷'),
        ('🔋', '動態加強器', '-', '本回合設施被移動過則輸出×2', '拆遷'),
        ('⚠️', '地基不穩定站', '-', '相鄰減益→翻倍；相鄰增益→-1', '拆遷'),
    ]),
    ('特殊設施', [
        ('💚', '嫉妒工廠', 'money→diamond', '金錢→1 鑽石，鑽石過商店×12', '特殊'),
        ('⬆️⬇️⬅️➡️', '物流方向', '-', '資源轉向指定方向', '特殊'),
        ('🍱', '員工食堂', 'goods→goods', '投入商品時額外獲得 1 人材', '特殊'),
        ('🏢', '人力派遣', 'money→money', '金錢投入時獲得 1 人材', '特殊'),
        ('🌐', '貿易代理', '-', '中央專屬。金錢→1%原料；商品→1%金錢', '特殊'),
        ('🔬', '科技研發', 'goods→goods', '中央專屬。商品投入時 +2%', '特殊'),
        ('🏛', '稅務局', '-', '每回合隨機設施升級，最終收益 -10%', '特殊'),
    ]),
]

for cat_name, items in categories:
    r = add_section(ws2, r, cat_name, 5)
    r = add_table_header(ws2, r, ['圖示', '名稱', '轉換', '效果說明', '分類'])
    for i, item in enumerate(items):
        r = add_table_row(ws2, r, item, i % 2 == 1)
    r += 1

# ══════════════════════════════════════
# Sheet 3: 合夥人一覽
# ══════════════════════════════════════
ws3 = wb.create_sheet()
setup_sheet(ws3, '合夥人一覽', [5, 12, 30, 30, 10])
r = 1
add_title_row(ws3, r, '合夥人一覽表', 5); r += 2

partner_cats = [
    ('惡魔系合夥人', [
        ('🪙', '貧窮神', '收益為零時 +bonus（逐次累積）', '獲得收益時 bonus 重置為 1', '惡魔'),
        ('👹', '暴食惡魔', '投入時原料⇌商品互換', '回合開始時原料/商品 -10%', '惡魔'),
        ('😴', '怠惰惡魔', '放棄選擇時 2/3 機率資源 +10', '1/3 機率資源變成 1', '惡魔'),
        ('💘', '慾望惡魔', '僅經過 2 設施時輸出×2', '經過 2+ 設施時資源÷設施數', '惡魔'),
        ('🔥', '激情惡魔', '減少效果被逆轉為增加', '超過 5 回合未遇減少→收益 -33%', '惡魔'),
        ('💚', '嫉妒惡魔', '獲得嫉妒工廠（金錢→鑽石→×12）', '非鑽石進入嫉妒工廠→收益 -50%', '惡魔'),
        ('💰', '貪婪惡魔', '每回合結束額外獲得收益 50%', '每輪開始目標 +50%', '惡魔'),
        ('👑', '傲慢惡魔', '惡魔負面效果不生效，每個惡魔 +10%', '無', '惡魔'),
        ('🕶️', '黑市商人', '每輪第一次商品→金錢×3', '之後每次倍率 -0.5（最低×1）', '惡魔'),
        ('🧨', '爆破工程師', '消耗 2 收益可主動摧毀設施', '每輪目標 +10%', '惡魔'),
    ]),
    ('基礎合夥人', [
        ('⛏️', '基礎原料商', '資源每次經過原料廠 +1 收益', '無', '基礎'),
        ('🏪', '基礎商店老闆', '資源每次經過商店系設施 +1 收益', '無', '基礎'),
        ('🏭', '基礎工廠主', '資源每次經過工廠 +1 收益', '無', '基礎'),
    ]),
    ('人力流合夥人', [
        ('📋', '人力仲介', '每回合開始 +1 人材', '每輪目標 +5%', '人力'),
        ('✊', '工會主席', '持有 3+ 人材時輸出 +50%', '少於 3 人材時 -20%', '人力'),
        ('🎩', '勞動部長', '使用人材後下次投入 +5', '未使用每回合 -5%', '人力'),
        ('👔', '人力資源總監', '一次全用人材時每個 +15', '每次只能使用 1 個人材', '人力'),
    ]),
    ('物流流合夥人', [
        ('🚢', '運輸大亨', '每次通過物流中心 +3', '未通過物流中心收益 -50%', '物流'),
        ('👑', '倉儲女王', '物流中心可疊加 2 層設施', '每個疊加行動費 +2', '物流'),
        ('🗺️', '路線規劃師', '每回合可免費移動物流中心', '物流 <2 個時輸出 -30%', '物流'),
        ('🚀', '快遞達人', '通過 4+ 格時 +10%', '少於 4 格收益減半', '物流'),
    ]),
    ('貿易流合夥人', [
        ('💹', '外匯交易員', '金錢↔商品轉換差值 10% 作收益', '原料時輸出 -50%', '貿易'),
        ('⚖️', '套利者', '同回合金→商→金時 +20%', '單向轉換 -20%', '貿易'),
        ('🎩', '壟斷者', '每 3 個商店系設施，商品→金錢 +5%', '少於 3 商店每少一個 -10%', '貿易'),
    ]),
    ('拆遷流合夥人', [
        ('🏚️', '地皮炒家', '移動設施後下次該位置 +20', '設施 >3 回合未移動 -1', '拆遷'),
        ('🌀', '混沌建築師', '每回合隨機移動設施到空格', '被移動時 -5% 收益', '拆遷'),
        ('💀', '廢墟掠奪者', '設施消滅後原位 +5', '廢墟 >3 格每格 -2', '拆遷'),
    ]),
    ('獨特合夥人', [
        ('👩‍💼', '譚雅', '每回合可用手牌與隨機 3 設施交換，所得 +1', '無', '獨特'),
        ('👩‍🔧', '蕾雅', '手牌同類設施升級 +1,+2%；排列時可合併', '無', '獨特'),
        ('🚛', '阿北，物流之王', '可將設施蓋在物流中心上', '無', '獨特'),
        ('🎖', '市長', '資源經過中央格子時最終收益 +5%', '無', '獨特'),
        ('🌟', '公路之星', '每回合中央 2 格隨機 +2', '未經中央收益減半', '獨特'),
        ('💣', '設施破壞者', '任何消滅設施時 +50 收益', '無', '獨特'),
        ('🌀', '場風大師', '每回合指定方向投入×2', '只能從指定方向投入', '獨特'),
        ('🏗', '大地主', '地圖擴大為 5×5', '設施補給時必須失去一個設施', '獨特'),
        ('🔨', '拆遷隊', '每 3 回合獲得免費重排機會（可累積）', '無', '獨特'),
    ]),
]

for cat_name, items in partner_cats:
    r = add_section(ws3, r, cat_name, 5)
    r = add_table_header(ws3, r, ['圖示', '名稱', '正面效果', '負面效果', '分類'])
    for i, item in enumerate(items):
        r = add_table_row(ws3, r, item, i % 2 == 1)
    r += 1

# ══════════════════════════════════════
# Sheet 4: 市場事件
# ══════════════════════════════════════
ws4 = wb.create_sheet()
setup_sheet(ws4, '市場事件', [5, 14, 40, 15, 12])
r = 1
add_title_row(ws4, r, '市場事件一覽（每 3 回合觸發，共 17 種）', 5); r += 2

r = add_text(ws4, r, 1, '事件規則：每 3 回合（第 3、6、9 回合）觸發一次。提前預告，觸發前 1 回合以藍色高亮受影響格子。', SMALL_FONT, 4)
r += 1

events = [
    ('🏗', '設施補給', '從隨機 3 個設施中選 1 加入手牌', '正面', '無'),
    ('📉', '原料大降', '金錢→原料×2，原料→金錢÷2，原料→商品×2', '混合', '無'),
    ('📈', '原料出口熱', '原料轉換效果×2（累積每次 +2x）', '正面', '無'),
    ('🛒', '商品熱銷', '商品→金錢額外 +2（累積每次 +2）', '正面', '無'),
    ('🌀', '颱風來襲', '隨機方向限制，物流失效，原料→金錢翻倍', '混合', '無'),
    ('⚔', '地區叛亂', '消滅 1~4 個角落設施', '負面', '角落格'),
    ('🌍', '地震', '所有設施往隨機方向滑動', '負面', '所有設施格'),
    ('☢', '危險廢棄物', '隨機空格放置 1 個爆破裝置（3 回合倒數）', '負面', '目標空格'),
    ('🎁', '譚雅的禮物', '手牌資源無視單位 +10', '正面', '無'),
    ('🎁', '蕾雅的禮物', '隨機一個設施永久 +5', '正面', '目標設施格'),
    ('🚫', '運輸異常', '物流失效 + 原料無法→商品', '負面', '無'),
    ('🍽', '重大食安', '員工食堂失效 + 金錢→商品時 -5', '負面', '無'),
    ('📋', '勞工保險', '最終收益扣除 人材數×2%', '負面', '無'),
    ('📈', '行業熱潮', '隨機一行 +10%，另一行 -10%', '混合', '兩行格子'),
    ('📊', '區域效應', '隨機一列 +10%，另一列 -10%', '混合', '兩列格子'),
    ('🌀', '地塊共鳴', '隨機 2×2 +10%，另一 2×2 -10%', '混合', '兩個 2×2'),
    ('😴', '平靜的一天', '什麼都沒發生', '中性', '無'),
]

r = add_table_header(ws4, r, ['圖示', '事件名稱', '效果說明', '性質', '預告高亮'])
for i, ev in enumerate(events):
    r = add_table_row(ws4, r, ev, i % 2 == 1)

# ══════════════════════════════════════
# Sheet 5: 商業行動
# ══════════════════════════════════════
ws5 = wb.create_sheet()
setup_sheet(ws5, '商業行動', [5, 14, 35, 10, 20])
r = 1
add_title_row(ws5, r, '商業行動一覽（每回合開始隨機 2 選 1）', 5); r += 2

r = add_text(ws5, r, 1, '費用公式：ceil(基礎費用 × max(1, 1 + (輪數-1) × 0.5))，費用從收益扣除', SMALL_FONT, 4)
r += 1

actions = [
    ('🏗', '購買設施', '從設施池隨機取 3 選 1 加入手牌', '4', '1/3 機率取得複合設施'),
    ('⚡', '觸發隨機事件', '立即觸發一個隨機市場事件', '2', ''),
    ('🔧', '重新排列設施', '自由拖曳移動所有設施，完成後鎖定', '2', '蕾雅可合併同類設施'),
    ('🤝', '招募合夥人', '隨機 3 選 1 招募合夥人', '5', ''),
    ('⬆', '提升設施數值', '選一個設施永久 +1 輸出', '3', ''),
    ('✨', '下次輸出加成', '下次元素以特定類型輸出時 +3', '3', ''),
    ('🔄', '更換預告事件', '重新抽一個下回合的隨機事件', '1', ''),
    ('😴', '什麼也沒發生', '跳過行動', '0', '怠惰惡魔可觸發'),
]

r = add_table_header(ws5, r, ['圖示', '行動名稱', '效果說明', '基礎費用', '備註'])
for i, a in enumerate(actions):
    r = add_table_row(ws5, r, a, i % 2 == 1)

# ══════════════════════════════════════
# Sheet 6: 特殊機制
# ══════════════════════════════════════
ws6 = wb.create_sheet()
setup_sheet(ws6, '特殊機制', [4, 16, 50])
r = 1
add_title_row(ws6, r, '特殊機制說明', 3); r += 2

mechanics = [
    ('廢墟系統', [
        ('設施消滅', '設施被消滅後，原位置變為廢墟（🏚️），資源經過不產生任何交互'),
        ('廢墟紀念碑', '只能放置在廢墟上，輸出 +5；消滅後還原為空地（不變廢墟）'),
        ('廢墟掠奪者', '合夥人效果：資源經過廢墟格 +5；廢墟超過 3 格時每格 -2 收益'),
    ]),
    ('複合設施', [
        ('取得方式', '購買設施時 1/3 機率取得兩格黏合卡牌'),
        ('放置規則', '需同時佔用相鄰兩格（方向固定），兩格皆須為空'),
        ('限制', '基礎設施、爆破裝置、嫉妒工廠等不會出現在複合設施中'),
    ]),
    ('百貨公司 2×2', [
        ('放置條件', '需要 2×2 區域的 4 個商店系設施'),
        ('佔用格子', '取代 4 個商店，佔用 2×2 格（1 錨點 + 3 部分格）'),
        ('視為商店', '計為 4 個商店（壟斷者等效果）'),
        ('消滅', '任一格被消滅時整個 2×2 變為廢墟'),
    ]),
    ('人材系統', [
        ('來源', '人力訓練中心（每回合+1）、人才市場、派遣總部、員工食堂等'),
        ('用途', '拖曳到設施格→本回合 +1；拖曳到元素卡→多投入一次；集體罷工台'),
        ('合夥人互動', '勞動轉換站消耗 2 人材×2、工會主席 3+ 人材→+50%'),
    ]),
    ('蕾雅升級', [
        ('手牌升級', '將手牌相同設施蓋在小鎮設施上 → +1 升級, +2% 永久百分比加成'),
        ('排列合併', '重新排列設施時可拖曳同類設施合併 → 升級值與百分比合併累加'),
        ('百分比加成', 'leyaPctMods 獨立於事件的 cellPctMods，永久生效不會被清除'),
    ]),
    ('動態難度', [
        ('追蹤指標', '每輪達標所需回合數、收益超額比率'),
        ('調整範圍', '倍率 0.6~2.5，影響下一輪目標'),
        ('重置', '失敗重開時難度重置為 1.0'),
    ]),
]

for mech_name, items in mechanics:
    r = add_section(ws6, r, mech_name, 3)
    for i, (label, desc) in enumerate(items):
        ws6.cell(row=r, column=1).font = BODY_FONT
        ws6.cell(row=r, column=2, value=label).font = H2_FONT
        ws6.cell(row=r, column=3, value=desc).font = BODY_FONT
        ws6.cell(row=r, column=3).alignment = WRAP
        if i % 2 == 1:
            for col in range(1, 4):
                ws6.cell(row=r, column=col).fill = ALT_ROW_FILL
        r += 1
    r += 1

# ══════════════════════════════════════
# Save
# ══════════════════════════════════════
output_path = r'E:\VT\VentureTown_企劃書.xlsx'
wb.save(output_path)
print(f'Done: {output_path}')
